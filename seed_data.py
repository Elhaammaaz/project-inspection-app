"""
Seed database with Building Assessment Report (BCAR) data from Excel
This loads all pre-defined inspection items, compliance items, tests, and CAPA records
"""

from openpyxl import load_workbook
import pandas as pd
from models import db, AssessmentItem, ComplianceItem, TestRecord, CAPARecord, \
                    SystemWeight, LookupTable, Project
from datetime import datetime


class BCAREXCELSeeder:
    
    def __init__(self, excel_file_path):
        self.excel_file = excel_file_path
        self.wb = load_workbook(excel_file_path)
    
    def seed_assessment_items(self, project_id):
        """Load all 694 assessment items from Building Assessment sheet"""
        
        print(f"Seeding assessment items for project {project_id}...")
        
        ws = self.wb['Building Assessment']
        
        # Skip header (row 1) and process data rows
        items_added = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                system = ws.cell(row=row_num, column=1).value
                subsystem = ws.cell(row=row_num, column=2).value
                component = ws.cell(row=row_num, column=3).value
                item_code = ws.cell(row=row_num, column=4).value
                inspection_item = ws.cell(row=row_num, column=5).value
                criteria = ws.cell(row=row_num, column=6).value
                test_method = ws.cell(row=row_num, column=7).value
                asset_tag = ws.cell(row=row_num, column=8).value
                snag_location = ws.cell(row=row_num, column=9).value
                snag_evidence_ref = ws.cell(row=row_num, column=10).value
                evidence_type = ws.cell(row=row_num, column=11).value
                rate = ws.cell(row=row_num, column=12).value
                # Column 13 is Score % (calculated)
                item_weight = ws.cell(row=row_num, column=14).value
                # Column 15 is Weighted Score (calculated)
                risk_criticality = ws.cell(row=row_num, column=16).value
                responsibility = ws.cell(row=row_num, column=17).value
                capa_id = ws.cell(row=row_num, column=18).value
                priority = ws.cell(row=row_num, column=19).value
                due_date = ws.cell(row=row_num, column=20).value
                status = ws.cell(row=row_num, column=21).value
                remarks = ws.cell(row=row_num, column=22).value
                
                # Skip if no data
                if not item_code:
                    continue
                
                # Check if already exists
                existing = AssessmentItem.query.filter_by(
                    project_id=project_id,
                    item_code=item_code
                ).first()
                
                if existing:
                    continue
                
                # Convert types
                rate = int(rate) if rate and isinstance(rate, (int, float)) else 3
                item_weight = float(item_weight) if item_weight and isinstance(item_weight, (int, float)) else 2.5
                risk_criticality = int(risk_criticality) if risk_criticality and isinstance(risk_criticality, (int, float)) else 0
                
                item = AssessmentItem(
                    project_id=project_id,
                    system=str(system).strip() if system else 'Unknown',
                    subsystem=str(subsystem).strip() if subsystem else None,
                    component=str(component).strip() if component else None,
                    item_code=str(item_code).strip(),
                    inspection_item=str(inspection_item).strip() if inspection_item else 'N/A',
                    criteria=str(criteria).strip() if criteria else None,
                    test_method=str(test_method).strip() if test_method else None,
                    asset_tag=str(asset_tag).strip() if asset_tag else None,
                    snag_location=str(snag_location).strip() if snag_location else None,
                    snag_evidence_ref=str(snag_evidence_ref).strip() if snag_evidence_ref else None,
                    evidence_type=str(evidence_type).strip() if evidence_type else None,
                    rate=rate,
                    item_weight=item_weight,
                    risk_criticality=risk_criticality,
                    responsibility=str(responsibility).strip() if responsibility else None,
                    capa_id=str(capa_id).strip() if capa_id else None,
                    priority=str(priority).strip() if priority else 'P3',
                    due_date=due_date if isinstance(due_date, (datetime, type(None))) else None,
                    status=str(status).strip() if status else 'Open',
                    remarks=str(remarks).strip() if remarks else None
                )
                
                db.session.add(item)
                items_added += 1
                
                # Commit every 100 items
                if items_added % 100 == 0:
                    db.session.commit()
                    print(f"  Added {items_added} items...")
            
            except Exception as e:
                print(f"  Error at row {row_num}: {str(e)}")
                continue
        
        db.session.commit()
        print(f"✓ Seeded {items_added} assessment items")
        return items_added
    
    def seed_compliance_items(self, project_id):
        """Load compliance items from Government Compliance Checklist sheet"""
        
        print(f"Seeding compliance items for project {project_id}...")
        
        ws = self.wb['Government Compliance Checklist']
        
        items_added = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                area = ws.cell(row=row_num, column=1).value
                description = ws.cell(row=row_num, column=2).value
                
                if not area:
                    continue
                
                # Check if already exists
                existing = ComplianceItem.query.filter_by(
                    project_id=project_id,
                    compliance_area=str(area).strip()
                ).first()
                
                if existing:
                    continue
                
                item = ComplianceItem(
                    project_id=project_id,
                    compliance_area=str(area).strip(),
                    description=str(description).strip() if description else None,
                    status='Open'
                )
                
                db.session.add(item)
                items_added += 1
            
            except Exception as e:
                print(f"  Error at row {row_num}: {str(e)}")
                continue
        
        db.session.commit()
        print(f"✓ Seeded {items_added} compliance items")
        return items_added
    
    def seed_test_records(self, project_id):
        """Load test register items"""
        
        print(f"Seeding test records for project {project_id}...")
        
        ws = self.wb['Test Register']
        
        items_added = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                test_code = ws.cell(row=row_num, column=1).value
                test_desc = ws.cell(row=row_num, column=2).value
                system = ws.cell(row=row_num, column=3).value
                
                if not test_code:
                    continue
                
                item = TestRecord(
                    project_id=project_id,
                    test_code=str(test_code).strip(),
                    test_description=str(test_desc).strip() if test_desc else None,
                    system=str(system).strip() if system else None,
                    test_result='Pending'
                )
                
                db.session.add(item)
                items_added += 1
            
            except Exception as e:
                print(f"  Error at row {row_num}: {str(e)}")
                continue
        
        db.session.commit()
        print(f"✓ Seeded {items_added} test records")
        return items_added
    
    def seed_capa_records(self, project_id):
        """Load CAPA register items"""
        
        print(f"Seeding CAPA records for project {project_id}...")
        
        ws = self.wb['CAPA Register']
        
        items_added = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                capa_id = ws.cell(row=row_num, column=1).value
                finding = ws.cell(row=row_num, column=2).value
                root_cause = ws.cell(row=row_num, column=3).value
                
                if not capa_id:
                    continue
                
                # Check if already exists
                existing = CAPARecord.query.filter_by(capa_id=str(capa_id).strip()).first()
                if existing:
                    continue
                
                item = CAPARecord(
                    project_id=project_id,
                    capa_id=str(capa_id).strip(),
                    finding_description=str(finding).strip() if finding else 'N/A',
                    root_cause=str(root_cause).strip() if root_cause else None,
                    status='Open'
                )
                
                db.session.add(item)
                items_added += 1
            
            except Exception as e:
                print(f"  Error at row {row_num}: {str(e)}")
                continue
        
        db.session.commit()
        print(f"✓ Seeded {items_added} CAPA records")
        return items_added
    
    def seed_system_weights(self, project_id):
        """Initialize system weights for all systems in the project"""
        
        print(f"Initializing system weights for project {project_id}...")
        
        # Get unique systems
        systems = db.session.query(AssessmentItem.system).filter_by(
            project_id=project_id
        ).distinct().all()
        
        weights_added = 0
        
        for (system,) in systems:
            existing = SystemWeight.query.filter_by(
                project_id=project_id,
                system_name=system
            ).first()
            
            if not existing:
                weight = SystemWeight(
                    project_id=project_id,
                    system_name=system,
                    weight=1.0  # Default weight
                )
                db.session.add(weight)
                weights_added += 1
        
        db.session.commit()
        print(f"✓ Initialized {weights_added} system weights")
        return weights_added
    
    def seed_lookups(self):
        """Load reference data from Lists sheet"""
        
        print("Seeding lookup tables...")
        
        ws = self.wb['Lists']
        
        lookups_added = 0
        
        # Priority values (Column 1)
        for row in range(2, 10):
            val = ws.cell(row=row, column=1).value
            if val:
                existing = LookupTable.query.filter_by(category='priority', value=str(val)).first()
                if not existing:
                    lookup = LookupTable(
                        category='priority',
                        value=str(val).strip(),
                        display_name=str(val).strip(),
                        sort_order=row-2
                    )
                    db.session.add(lookup)
                    lookups_added += 1
        
        # Status values (Column 2)
        for row in range(2, 10):
            val = ws.cell(row=row, column=2).value
            if val:
                existing = LookupTable.query.filter_by(category='status', value=str(val)).first()
                if not existing:
                    lookup = LookupTable(
                        category='status',
                        value=str(val).strip(),
                        display_name=str(val).strip(),
                        sort_order=row-2
                    )
                    db.session.add(lookup)
                    lookups_added += 1
        
        # Responsibility values (Column 3)
        for row in range(2, 10):
            val = ws.cell(row=row, column=3).value
            if val:
                existing = LookupTable.query.filter_by(category='responsibility', value=str(val)).first()
                if not existing:
                    lookup = LookupTable(
                        category='responsibility',
                        value=str(val).strip(),
                        display_name=str(val).strip(),
                        sort_order=row-2
                    )
                    db.session.add(lookup)
                    lookups_added += 1
        
        db.session.commit()
        print(f"✓ Seeded {lookups_added} lookup entries")
        return lookups_added
    
    def seed_all(self, project_id):
        """Seed all data for a project"""
        
        print("\n" + "=" * 60)
        print("SEEDING DATABASE WITH BCAR EXCEL DATA")
        print("=" * 60)
        
        self.seed_assessment_items(project_id)
        self.seed_compliance_items(project_id)
        self.seed_test_records(project_id)
        self.seed_capa_records(project_id)
        self.seed_system_weights(project_id)
        self.seed_lookups()
        
        print("\n" + "=" * 60)
        print("✓ DATABASE SEEDING COMPLETE")
        print("=" * 60 + "\n")


# Usage:
# from flask import create_app
# from config import Config
# app = create_app()
# with app.app_context():
#     seeder = BCAREXCELSeeder('Building Assessment Report (BCAR) – v12.xlsx')
#     seeder.seed_all(project_id)
